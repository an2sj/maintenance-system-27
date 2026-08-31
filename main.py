# -*- coding: utf-8 -*-
"""
نظام إدارة الصيانة وأوامر العمل
مستشفى الأمير محمد بن ناصر | تجمع جازان الصحي

تطبيق FastAPI + SQLite جاهز للنشر السحابي (Render / Replit):
  - يستمع تلقائياً على 0.0.0.0:$PORT (المنفذ الديناميكي الذي تحدده المنصة)
  - قاعدة بيانات SQLite محلية تُحفظ بين عمليات إعادة النشر عبر قرص دائم
  - حماية لوحة التحكم بكلمة مرور (الصفحات العامة تبقى مفتوحة)
  - واجهة ثنائية اللغة (عربي/إنجليزي)
  - مرفقات ميديا (صور/فيديو) + رقم طلب فريد تلقائي
  - متجاوب بالكامل مع الآيباد والجوال
"""

import hashlib
import io
import json
import os
import re
import secrets
import shutil
import smtplib
import urllib.error
import urllib.request
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote, unquote
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import qrcode
import uvicorn
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# --------------------------------------------------------------------------- #
# الإعدادات والثوابت
# --------------------------------------------------------------------------- #
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"

# مسار قاعدة البيانات: يسمح للبيئة السحابية بتوجيهه إلى قرص دائم
# على Render: نستخدم القرص المثبت على /data تلقائياً إن توفر، وإلا نستخدم مجلد المشروع محلياً
_default_db = Path("/data/maintenance.db") if Path("/data").is_dir() else BASE_DIR / "data" / "maintenance.db"
DB_PATH = Path(os.environ.get("DB_PATH", str(_default_db)))

# المنطقة الزمنية (افتراضية: الرياض)
try:
    TZ = ZoneInfo(os.environ.get("APP_TIMEZONE", "Asia/Riyadh"))
except ZoneInfoNotFoundError:
    TZ = ZoneInfo("UTC")

# هوية المؤسسة (بدون شركة مقاولات)
HOSPITAL_NAME = os.environ.get("HOSPITAL_NAME", "مستشفى الأمير محمد بن ناصر")
ORG_RIGHT = os.environ.get("ORG_RIGHT", "تجمع جازان الصحي")
ORG_LEFT = os.environ.get("ORG_LEFT", "")  # أُزيلت شركة المقاولات نهائياً
DEPARTMENT_NAME = os.environ.get("DEPARTMENT_NAME", "إدارة الصيانة والعمليات")

# شعار تجمع جازان الصحي (يُستخدم في الهيدر وفي مركز رمز QR)
LOGO_PATH = STATIC_DIR / "logo.jpg"
if not LOGO_PATH.exists():
    LOGO_PATH = STATIC_DIR / "logo.png"
    if not LOGO_PATH.exists():
        LOGO_PATH = None

# --------------------------------------------------------------------------- #
# مناطق التخزين الدائم (الميديا)
# --------------------------------------------------------------------------- #
# تُخزَّن المرفقات على القرص الدائم (مجلد البيانات) حتى لا تُفقد عند إعادة النشر
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", DB_PATH.parent / "uploads"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MEDIA_URL_PREFIX = "/media"
ALLOWED_MEDIA = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".gif": "image/gif", ".webp": "image/webp",
    ".mp4": "video/mp4", ".mov": "video/quicktime", ".webm": "video/webm",
}
MAX_MEDIA = 5  # الحد الأقصى لعدد المرفقات في بلاغ واحد

# --------------------------------------------------------------------------- #
# أنواع الصيانة ونوع المشكلة الديناميكي
# --------------------------------------------------------------------------- #
MAINTENANCE_TYPES = ["كهرباء", "تكييف", "سباكة", "مدني", "نجارة", "أمن وسلامة", "أخرى"]

PROBLEM_TYPES = {
    "كهرباء": ["انقطاع التيار", "تلف منفذ/مقبس", "مشكلة إضاءة", "قاطع كهربائي", "مروحة/شفاط", "أخرى"],
    "تكييف": ["لا يبرد", "تسريب ماء", "ضجيج", "لا يعمل", "ريموت كونترول", "أخرى"],
    "سباكة": ["تسريب ماء", "انسداد", "حرارة/ماء ساخن", "صرف صحي", "خلاط/صنابير", "أخرى"],
    "مدني": ["تشقق/تصدع", "تسرب سقف", "بلاط/سيراميك", "دهانات", "أبواب/نوافذ", "أخرى"],
    "نجارة": ["خزانة/دولاب", "كسر أثاث", "طاولة/منضدة", "تشليح خشب", "كرسي/مقعد", "أخرى"],
    "أمن وسلامة": ["طفاية حريق", "نظام إنذار", "مخرج طوارئ", "إضاءة طوارئ", "بوابة/حاجز", "أخرى"],
    "أخرى": ["أخرى"],
}

PRIORITIES = ["عادية", "مستعجلة", "طارئة"]

STATUS_META = {
    "pending": ("بانتظار الاعتماد", "warn"),
    "approved": ("معتمد - قيد التنفيذ", "info"),
    "done": ("منجز ومغلق", "ok"),
    "rejected": ("مرفوض", "bad"),
}

MONTHS_AR = [
    "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
    "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر",
]
MONTHS_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
WEEKDAYS_AR = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]

# --------------------------------------------------------------------------- #
# المواقع / المباني
# --------------------------------------------------------------------------- #
PUBLIC_FACILITIES = [
    "المكتب", "الاستقبال", "الحديقة", "مواقف السيارات",
    "غرفة الأمن", "المضخات", "غرفة الكهرباء",
]

# مبنى T5 و T6: 6 أدوار × 4 شقق = 24 شقة (الدور 1..6)
T56_UNITS = []
for floor in range(1, 7):
    start = (floor - 1) * 4 + 1
    for i in range(4):
        T56_UNITS.append(f"الدور {floor} - شقة {start + i}")

BUILDINGS = ["المرافق العامة", "مبنى T5", "مبنى T6", "مبنى T1"]

# المصادقة الاختيارية
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "")
AUTH_COOKIE = "mn_maintenance_auth"
PROTECTED = bool(DASHBOARD_PASSWORD)


def auth_token() -> str:
    return hashlib.sha256(("mn-maintenance:" + DASHBOARD_PASSWORD).encode()).hexdigest()


def is_public_path(path: str) -> bool:
    return (
        path in ("/login", "/logout", "/healthz")
        or path.startswith(("/report", "/track/", "/qr", "/poster-qr", "/media", "/static"))
    )


# --------------------------------------------------------------------------- #
# اللغة (ثنائية اللغة: عربي / إنجليزي)
# --------------------------------------------------------------------------- #
LANG_COOKIE = "mn_lang"
SUPPORTED_LANGS = ("ar", "en")


def pick_lang(lang: str | None, cookie: str | None = None) -> str:
    """اختيار اللغة: أولوية المعامل ثم الكوكيز ثم العربية."""
    for candidate in (lang, cookie):
        if candidate in SUPPORTED_LANGS:
            return candidate
    return "ar"


# قواميس الترجمة الشاملة (عربي -> إنجليزي حسب المفتاح)
AR = {}
EN = {}


def I(key, ar, en):
    """تسجيل نص مترجم."""
    AR[key] = ar
    EN[key] = en


def tr(lang: str, key: str) -> str:
    table = EN if lang == "en" else AR
    return table.get(key, key)


def tr_value(lang: str, value: str, map_key: str = None) -> str:
    """ترجمة قيمة ديناميكية (مبنى/قسم/حالة/أولوية)."""
    if lang != "en" or not value:
        return value
    table = VALUE_EN[map_key] if map_key in VALUE_EN else {}
    return table.get(value, value)


VALUE_EN = {
    "status": {},
    "priority": {},
    "maintenance": {},
    "building": {},
}


def setup_values():
    for k, (ar, _c) in STATUS_META.items():
        VALUE_EN["status"][ar] = STATUS_META_EN.get(k, ar)
    for ar, en in PRIORITY_EN.items():
        VALUE_EN["priority"][ar] = en
    for ar, en in MAINTENANCE_EN.items():
        VALUE_EN["maintenance"][ar] = en
    for ar, en in BUILDING_EN.items():
        VALUE_EN["building"][ar] = en


STATUS_META_EN = {
    "pending": "Pending Approval",
    "approved": "Approved - In Progress",
    "done": "Completed & Closed",
    "rejected": "Rejected",
}
PRIORITY_EN = {"عادية": "Normal", "مستعجلة": "Urgent", "طارئة": "Critical"}
BUILDING_EN = {
    "المرافق العامة": "Public Facilities",
    "مبنى T5": "Building T5",
    "مبنى T6": "Building T6",
    "مبنى T1": "Building T1",
}
MAINTENANCE_EN = {
    "كهرباء": "Electrical", "تكييف": "AC / Cooling", "سباكة": "Plumbing",
    "مدني": "Civil", "نجارة": "Carpentry", "أمن وسلامة": "Safety",
    "أخرى": "Other",
}


# ======================================================================= #
# النصوص العامة (واجهة ثنائية اللغة كاملة)
# ======================================================================= #
I("nav_dash", "لوحة المتابعة", "Dashboard")
I("nav_new", "أمر عمل جديد", "New Work Order")
I("nav_incoming", "الطلبات الواردة", "Incoming Requests")
I("nav_monthly", "التقرير الشهري", "Monthly Report")
I("nav_poster", "ملصق QR", "QR Poster")
I("nav_logout", "خروج", "Logout")
I("lang_switch", "EN", "عربي")

I("dash_title", "لوحة متابعة أوامر العمل", "Work Orders Dashboard")
I("dash_totals", "إجمالي أوامر العمل", "Total Work Orders")
I("dash_pending_review", "بانتظار التدقيق", "Pending Review")
I("dash_in_progress", "قيد التنفيذ", "In Progress")
I("dash_done", "منجزة ومغلقة", "Completed & Closed")
I("dash_urgent", "أوامر طارئة", "Critical Orders")
I("search_ph", "بحث: الاسم / رقم الأمر / الموقع / الفني / الجوال...", "Search: name / order no / location / technician / phone...")
I("all_status", "كل الحالات", "All Statuses")
I("filter", "تصفية", "Filter")
I("clear", "مسح", "Clear")
I("col_order", "رقم الأمر", "Order No")
I("col_year", "السنة", "Year")
I("col_title", "التاريخ", "Date")
I("col_reporter", "اسم المبلّغ", "Reporter")
I("col_contact", "رقم التواصل", "Contact")
I("col_location", "الموقع", "Location")
I("col_section", "نوع الصيانة", "Maintenance")
I("col_problem", "نوع المشكلة", "Problem Type")
I("col_priority", "الأولوية", "Priority")
I("col_technician", "الفني", "Technician")
I("col_status", "الحالة", "Status")
I("col_source", "المصدر", "Source")
I("col_actions", "إجراءات", "Actions")
I("action_view", "عرض", "View")
I("action_edit", "تعديل", "Edit")

I("new_title", "إنشاء أمر عمل جديد", "Create New Work Order")
I("new_sub", "يُحفظ الأمر مباشرة ويظهر فوراً في لوحة المتابعة", "Saved instantly and shown in the dashboard")
I("f_reporter", "اسم المبلّغ", "Reporter Name")
I("f_contact", "رقم التواصل", "Contact Number")
I("f_maintenance_type", "نوع الصيانة", "Maintenance Type")
I("f_problem_type", "نوع المشكلة", "Problem Type")
I("f_priority", "الأولوية", "Priority")
I("f_building", "المبنى / الموقع", "Building / Location")
I("f_unit", "الوحدة / الشقة / التفصيل", "Unit / Apartment / Detail")
I("f_technician", "الفني المكلف", "Assigned Technician")
I("f_description", "وصف المشكلة", "Problem Description")
I("f_media", "مرفقات (صور / فيديو)", "Attachments (Photo / Video)")
I("f_media_hint", "يمكنك إرفاق حتى 5 صور أو مقطع فيديو لمساعدة الفني", "You can attach up to 5 photos or a video to help the technician")
I("f_location", "الموقع المفصل", "Detailed Location")
I("btn_save", "💾 حفظ وإصدار الأمر", "💾 Save & Issue Order")
I("btn_cancel", "إلغاء", "Cancel")
I("btn_submit", "📨 إرسال البلاغ", "📨 Submit Report")

I("title_report", "بلاغ صيانة", "Maintenance Report")
I("sub_report", "نموذج رفع بلاغ / طلب صيانة — يصل فوراً إلى قسم الصيانة للتدقيق والاعتماد", "Report / maintenance request form — reaches maintenance instantly for review & approval")
I("report_choose_issue", "اضغط على نوع الصيانة ثم اختر تفاصيل المشكلة", "Select maintenance type then choose the issue details")

I("thanks_title", "تم استلام بلاغك بنجاح", "Your report was received successfully")
I("thanks_ref", "رقم مرجعي لتتبع بلاغك", "Your reference number")
I("thanks_note", "سيدرّس قسم الصيانة البلاغ ويحوّله إلى أمر عمل رسمي خلال أقرب وقت.", "Maintenance will review and convert to an official work order soon.")
I("thanks_again", "تقديم بلاغ آخر", "Submit Another Report")

I("track_status", "حالة أمر العمل", "Work order status")
I("track_notfound", "رابط غير صالح", "Invalid link")
I("track_notfound_note", "هذا الرمز غير مرتبط بأمر عمل موجود.", "This code is not linked to an existing work order.")
I("track_inquiry", "للاستفسار يرجى التواصل مع قسم الصيانة", "For inquiries, contact the maintenance department")
I("track_assigned_soon", "سيتم الإسناد قريباً", "To be assigned soon")
I("f_completion", "تاريخ الإنجاز", "Completion Date")
I("f_received", "تاريخ الاستلام", "Received Date")

I("edit_title", "تعديل أمر العمل", "Edit Work Order")
I("btn_save_changes", "💾 حفظ التعديلات", "💾 Save Changes")
I("btn_update", "💾 حفظ التحديثات", "💾 Save Updates")
I("col_notes", "ملاحظات / الإجراء", "Notes / Action Taken")
I("btn_delete", "حذف الأمر", "Delete Order")
I("btn_print_header", "🖨 طباعة بالهيدر الرسمي", "🖨 Print with Official Header")
I("btn_track", "صفحة التتبع", "Tracking Page")
I("btn_print", "🖨 طباعة", "🖨 Print")

I("inc_title", "الطلبات الواردة — بانتظار التدقيق", "Incoming Requests — Pending Review")
I("inc_sub", "البلاغات المرفوعة عبر مسح رمز QR تظهر هنا أولاً ولا تدخل السجل الرسمي إلا بعد اعتمادها", "QR submissions appear here first and enter the official log only after approval")
I("btn_approve", "اعتماد ✓", "Approve ✓")
I("btn_reject", "رفض ✕", "Reject ✕")
I("btn_review", "تدقيق", "Review")
I("inc_empty_title", "لا توجد طلبات واردة جديدة", "No new incoming requests")
I("inc_empty_note", "عند مسح رمز QR الموجود على ملصق البلاغات سيظهر البلاغ هنا للتدقيق والاعتماد.", "Scanning the QR poster will bring reports here for review and approval.")

I("mth_title", "التقرير الشهري لأوامر العمل", "Monthly Work Orders Report")
I("mth_total", "إجمالي الأوامر", "Total Orders")
I("mth_done", "منجزة ومغلقة", "Completed & Closed")
I("mth_in_progress", "قيد التنفيذ", "In Progress")
I("mth_pending", "بانتظار الاعتماد", "Pending Approval")
I("mth_rejected", "مرفوضة", "Rejected")
I("mth_urgent", "أوامر طارئة", "Critical Orders")
I("mth_rate", "نسبة الإنجاز", "Completion Rate")
I("mth_by_section", "قسم الصيانة", "Section")
I("mth_by_priority", "الأولوية", "Priority")
I("mth_by_status", "الحالة", "Status")
I("mth_by_source", "المصدر", "Source")
I("mth_details", "تفاصيل أوامر العمل", "Work order details")
I("mth_show", "عرض", "Show")
I("mth_print", "🖨 طباعة التقرير", "🖨 Print Report")
I("mth_count", "العدد", "Count")

I("login_title", "لوحة التحكم", "Dashboard")
I("login_pass", "كلمة المرور", "Password")
I("login_btn", "تسجيل الدخول", "Sign In")
I("login_err", "كلمة المرور غير صحيحة. حاول مرة أخرى.", "Incorrect password. Try again.")
I("login_ph", "••••••••", "••••••••")

I("f_unit_placeholder", "مثال: الدور 3 - شقة 12", "e.g. Floor 3 - Apartment 12")
I("select_placeholder", "اختر...", "Select...")

setup_values()


# --------------------------------------------------------------------------- #
# قاعدة البيانات
# --------------------------------------------------------------------------- #
DB_PATH.parent.mkdir(parents=True, exist_ok=True)


@contextmanager
def db():
    conn = __import__("sqlite3").connect(DB_PATH)
    conn.row_factory = __import__("sqlite3").Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    with db() as c:
        c.executescript(
            """
            CREATE TABLE IF NOT EXISTS work_orders (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no      TEXT UNIQUE,
                reference_id  TEXT UNIQUE,
                reporter_name TEXT NOT NULL,
                contact       TEXT NOT NULL DEFAULT '',
                location      TEXT NOT NULL DEFAULT '',
                building      TEXT NOT NULL DEFAULT '',
                unit          TEXT NOT NULL DEFAULT '',
                section       TEXT NOT NULL DEFAULT '',
                problem_type  TEXT NOT NULL DEFAULT '',
                priority      TEXT NOT NULL DEFAULT 'عادية',
                technician    TEXT NOT NULL DEFAULT '',
                description   TEXT NOT NULL DEFAULT '',
                notes         TEXT NOT NULL DEFAULT '',
                media         TEXT NOT NULL DEFAULT '',
                status        TEXT NOT NULL DEFAULT 'pending',
                source        TEXT NOT NULL DEFAULT 'manual',
                token         TEXT UNIQUE,
                created_at    TEXT NOT NULL,
                completed_at  TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_wo_status ON work_orders(status);
            CREATE INDEX IF NOT EXISTS idx_wo_created ON work_orders(created_at);
            """
        )
        # ترحيل للأرشيف: إضافة الأعمدة الجديدة إن لم تكن موجودة (توافق بيانات قديمة)
        existing = {r[1] for r in c.execute("PRAGMA table_info(work_orders)").fetchall()}
        mig = {
            "reference_id": "TEXT",
            "building": "TEXT NOT NULL DEFAULT ''",
            "unit": "TEXT NOT NULL DEFAULT ''",
            "problem_type": "TEXT NOT NULL DEFAULT ''",
            "media": "TEXT NOT NULL DEFAULT ''",
        }
        for col, ddl in mig.items():
            if col not in existing:
                c.execute(f"ALTER TABLE work_orders ADD COLUMN {col} {ddl}")


init_db()


def now_local() -> datetime:
    return datetime.now(TZ).replace(tzinfo=None)


def fmt(dt: datetime, with_time: bool = True) -> str:
    s = dt.strftime("%Y-%m-%d")
    if with_time:
        s += " " + dt.strftime("%H:%M")
    return s


# --------------------------------------------------------------------------- #
# تطبيق FastAPI
# --------------------------------------------------------------------------- #
app = FastAPI(
    title=f"نظام أوامر العمل - {HOSPITAL_NAME}",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    openapi_url="/api/openapi.json",
)
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


def _load_meta(lang: str):
    return {
        k: (v[0], v[1]) for k, v in STATUS_META.items()
    }


templates.env.globals.update(
    HOSPITAL_NAME=HOSPITAL_NAME,
    ORG_RIGHT=ORG_RIGHT,
    ORG_LEFT=ORG_LEFT,
    DEPARTMENT_NAME=DEPARTMENT_NAME,
    LOGO=("/static/" + LOGO_PATH.name) if LOGO_PATH else None,
    MAINTENANCE_TYPES=MAINTENANCE_TYPES,
    PROBLEM_TYPES=PROBLEM_TYPES,
    PUBLIC_FACILITIES=PUBLIC_FACILITIES,
    T56_UNITS=T56_UNITS,
    BUILDINGS=BUILDINGS,
    PRIORITIES=PRIORITIES,
    STATUS_META=STATUS_META,
    MONTHS_AR=MONTHS_AR,
    MONTHS_EN=MONTHS_EN,
)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# وسم تحويل التاريخ للنماذج
def dt_filter(value, with_time: bool = True):
    if not value:
        return "—"
    try:
        return fmt(datetime.strptime(str(value), "%Y-%m-%d %H:%M"), with_time)
    except Exception:
        try:
            return fmt(datetime.strptime(str(value), "%Y-%m-%d"), with_time)
        except Exception:
            return str(value)


templates.env.filters["dt"] = dt_filter


# --------------------------------------------------------------------------- #
# المصادقة
# --------------------------------------------------------------------------- #
@app.middleware("http")
async def auth_gate(request: Request, call_next):
    if PROTECTED and not is_public_path(request.url.path):
        if request.cookies.get(AUTH_COOKIE) != auth_token():
            nxt = request.url.path
            if request.url.query:
                nxt += "?" + request.url.query
            return RedirectResponse(f"/login?next={nxt}", status_code=303)
    return await call_next(request)


@app.middleware("http")
async def lang_gate(request: Request, call_next):
    """تحديد اللغة الحالية من المعامل ?lang= أو من الكوكيز، وتخزينها في حالة الطلب."""
    request.state.lang = pick_lang(request.query_params.get("lang"), request.cookies.get(LANG_COOKIE))
    response = await call_next(request)
    # حفظ تفضيل اللغة في الكوكيز ليستمر عبر الصفحات
    if request.query_params.get("lang"):
        response.set_cookie(
            LANG_COOKIE, request.state.lang,
            max_age=60 * 60 * 24 * 365, samesite="lax",
        )
    return response


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, next: str = "/", err: str = ""):
    return templates.TemplateResponse(
        request, "login.html",
        context(request, next=next or "/", err=err),
    )


@app.post("/login")
def login_submit(password: str = Form(""), next: str = Form("/")):
    if password != DASHBOARD_PASSWORD:
        return RedirectResponse(f"/login?err=1&next={next}", status_code=303)
    safe_next = next if next.startswith("/") and not next.startswith("//") else "/"
    resp = RedirectResponse(safe_next, status_code=303)
    resp.set_cookie(
        AUTH_COOKIE,
        auth_token(),
        max_age=60 * 60 * 24 * 30,
        httponly=True,
        samesite="lax",
        secure=os.environ.get("COOKIE_SECURE", "1") == "1",
    )
    return resp


@app.get("/logout")
def logout():
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie(AUTH_COOKIE)
    return resp


# --------------------------------------------------------------------------- #
# دوال مساعدة
# --------------------------------------------------------------------------- #
FLASH_MESSAGES = {
    "created": ("تم إنشاء أمر العمل بنجاح.", "تم إنشاء أمر العمل بنجاح.", "ok"),
    "updated": ("تم تحديث بيانات أمر العمل.", "تم تحديث بيانات أمر العمل.", "ok"),
    "deleted": ("تم حذف أمر العمل.", "تم حذف أمر العمل.", "ok"),
    "approved": ("تم اعتماد الطلب وتحويله إلى أمر عمل.", "تم اعتماد الطلب وتحويله إلى أمر عمل.", "ok"),
    "rejected": ("تم رفض الطلب.", "تم رفض الطلب.", "warn"),
    "err": ("يرجى تعبئة جميع الحقول الإلزامية.", "يرجى تعبئة جميع الحقول الإلزامية.", "bad"),
}

def context(request: Request, **extra) -> dict:
    with db() as c:
        pending_count = c.execute(
            "SELECT COUNT(*) FROM work_orders WHERE status='pending'"
        ).fetchone()[0]
    lang = getattr(request.state, "lang", "ar")
    # جمع إشعارات الوميض من معاملات الاستعلام (created=1, deleted=1, err=1 ...)
    flashes = []
    for key, qv in request.query_params.items():
        if key in FLASH_MESSAGES and qv in ("1",):
            ar, en, level = FLASH_MESSAGES[key]
            flashes.append((level, en if lang == "en" else ar))
    data = {
        "request": request,
        "path": request.url.path,
        "pending_count": pending_count,
        "protected": PROTECTED,
        "flashes": flashes,
        "now": now_local(),
        "lang": lang,
        "LANG": lang,
        "tr": lambda key: tr(lang, key),
        "trv": lambda value, mk=None: tr_value(lang, value, mk),
        "T_STATUS": lambda v: tr_value(lang, v, "status"),
        "T_PRIO": lambda v: tr_value(lang, v, "priority"),
        "T_MAINT": lambda v: tr_value(lang, v, "maintenance"),
        "T_BUILD": lambda v: tr_value(lang, v, "building"),
    }
    data.update(extra)
    return data


def base_url(request: Request) -> str:
    env = os.environ.get("PUBLIC_BASE_URL", "").strip().rstrip("/")
    if env:
        return env
    return str(request.base_url).rstrip("/")


def make_qr_png(url: str, with_logo: bool = True) -> Response:
    qr = qrcode.QRCode(border=2, box_size=10, error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#0f172a", back_color="white").convert("RGB")
    # دمج شعار تجمع جازان في مركز رمز QR
    if with_logo and LOGO_PATH:
        try:
            from PIL import Image
            logo = Image.open(LOGO_PATH).convert("RGBA")
            logo = logo.resize((70, 70), Image.LANCZOS)
            # خلفية بيضاء خلف الشعار لإبراز الوضوح
            bg = Image.new("RGBA", logo.size, (255, 255, 255, 255))
            base = img
            # مركز الصورة
            w, h = img.size
            cx, cy = (w - logo.width) // 2, (h - logo.height) // 2
            bg.paste(logo, (0, 0), logo)
            base.paste(bg, (cx, cy), bg)
            bio = io.BytesIO()
            base.save(bio, format="PNG")
            return Response(content=bio.getvalue(), media_type="image/png")
        except Exception:
            pass
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    return Response(content=bio.getvalue(), media_type="image/png")


def save_media(files) -> list[str]:
    """حفظ الملفات المرفوعة على القرص الدائم، وإرجاع أسماء الملفات النسبية."""
    saved = []
    for f in (files or []):
        if not f or not f.filename:
            continue
        # التحقق من النوع والسلامة
        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_MEDIA:
            continue
        if len(saved) >= MAX_MEDIA:
            break
        # اسم ملف آمن وفريد
        safe_name = secrets.token_hex(6) + ext
        dest = UPLOAD_DIR / safe_name
        try:
            with dest.open("wb") as out:
                shutil.copyfileobj(f.file, out)
            saved.append(f"/{MEDIA_URL_PREFIX.lstrip('/')}/{safe_name}")
        except Exception:
            continue
    return saved


# دالة إرجاع قوائم المرفقات بصيغة آمنة
def media_list(media_str: str) -> list[str]:
    if not media_str:
        return []
    return [p.strip() for p in media_str.split(",") if p.strip()]


# ======================================================================= #
# التصدير (PDF / Excel) للتقرير الشهري + إشعارات البريد/WhatsApp
# ======================================================================= #

# خط عربي مفتوح المصدر (Noto Naskh Arabic) مضمّن مع المشروع
PDF_FONT_PATH = STATIC_DIR / "fonts" / "NotoNaskhArabic.ttf"


def _pdf_register_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    try:
        pdfmetrics.registerFont(TTFont("Naskh", str(PDF_FONT_PATH)))
        pdfmetrics.registerFontFamily("Naskh", normal="Naskh", bold="Naskh",
                                      italic="Naskh", boldItalic="Naskh")
        return "Naskh"
    except Exception:
        return "Helvetica"


_FONT_NAME = _pdf_register_fonts()


def _ar_shape(text) -> str:
    """إعادة تشكيل النص العربي وإعادة ترتيبه ليُعرض صحيحاً داخل reportlab."""
    t = str(text or "")
    if not t:
        return ""
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(t))
    except Exception:
        return t


def _pdf_escape(text) -> str:
    return (str(text or "").replace("&", "&amp;")
            .replace("<", "&lt;").replace(">", "&gt;"))


def _pdf_cell(text, bold=False, size=8):
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph
    val = _pdf_escape(_ar_shape(text))
    if bold:
        val = f"<b>{val}</b>"
    return Paragraph(val, ParagraphStyle(
        "c", fontName=_FONT_NAME, fontSize=size, leading=size + 3,
        wordWrap="CJK",
    ))


def monthly_pdf_bytes(rows, totals, year, month, by_section) -> bytes:
    """بناء ملف PDF للتقرير الشهري وإرجاع وحدات البايت."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=11 * mm, rightMargin=11 * mm,
        topMargin=11 * mm, bottomMargin=13 * mm,
        title=f"التقرير الشهري {MONTHS_AR[month - 1]} {year}",
    )
    title = ParagraphStyle("t", fontName=_FONT_NAME, fontSize=15, leading=20,
                           alignment=TA_CENTER)
    sub = ParagraphStyle("s", fontName=_FONT_NAME, fontSize=8, leading=11,
                         alignment=TA_CENTER, textColor=colors.HexColor("#64748b"))
    secl = ParagraphStyle("sec", fontName=_FONT_NAME, fontSize=11, leading=15,
                          spaceBefore=6, spaceAfter=4)

    total = totals.get("total", len(rows))
    story = []

    lh_cell = Paragraph(
        f"<b>{_pdf_escape(_ar_shape(HOSPITAL_NAME))}</b><br/>"
        f"{_pdf_escape(_ar_shape(ORG_RIGHT))} · {_pdf_escape(_ar_shape(DEPARTMENT_NAME))}",
        ParagraphStyle("lhc", fontName=_FONT_NAME, fontSize=11, leading=15,
                       alignment=TA_CENTER),
    )
    if LOGO_PATH:
        story.append(Table(
            [[Image(str(LOGO_PATH), width=22 * mm, height=22 * mm), lh_cell,
              Image(str(LOGO_PATH), width=22 * mm, height=22 * mm)]],
            colWidths=[28 * mm, 144 * mm, 28 * mm],
        ))
    else:
        story.append(Table([[lh_cell]], colWidths=[200 * mm]))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph(_pdf_escape(_ar_shape(
        f"التقرير الشهري لأوامر العمل — {MONTHS_AR[month - 1]} {year}")), title))
    story.append(Paragraph(_pdf_escape(_ar_shape("MONTHLY MAINTENANCE REPORT")), sub))
    story.append(Spacer(1, 3 * mm))

    kpi = Table(
        [[_pdf_cell("إجمالي الأوامر", True), _pdf_cell("منجزة", True),
          _pdf_cell("قيد التنفيذ", True), _pdf_cell("بانتظار الاعتماد", True),
          _pdf_cell("مرفوضة", True), _pdf_cell("نسبة الإنجاز", True)],
         [_ar_shape(str(total)), _ar_shape(str(totals.get("done", 0))),
          _ar_shape(str(totals.get("approved", 0))),
          _ar_shape(str(totals.get("pending", 0))),
          _ar_shape(str(totals.get("rejected", 0))),
          _ar_shape(f"{totals.get('rate', 0)}%")]],
        colWidths=[34 * mm] * 6,
    )
    kpi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.white, colors.HexColor("#f0fdfa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(kpi)

    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(_pdf_escape(_ar_shape("توزيع البلاغات حسب القسم")), secl))
    dist_rows = [[_pdf_cell("القسم / Type", True), _pdf_cell("العدد", True),
                  _pdf_cell("النسبة", True)]]
    for k, v in sorted(by_section.items(), key=lambda kv: -kv[1]):
        dist_rows.append([
            _pdf_cell(k), _ar_shape(str(v)),
            _ar_shape(f"{round(v * 100 / total) if total else 0}%"),
        ])
    dt = Table(dist_rows, colWidths=[120 * mm, 40 * mm, 40 * mm])
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    story.append(dt)

    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(_pdf_escape(_ar_shape("تفاصيل أوامر العمل")), secl))
    header = [_pdf_cell("#", True), _pdf_cell("رقم الأمر", True),
              _pdf_cell("التاريخ", True), _pdf_cell("المبلّغ", True),
              _pdf_cell("الموقع", True), _pdf_cell("القسم", True),
              _pdf_cell("الفني", True), _pdf_cell("الأولوية", True),
              _pdf_cell("الحالة", True)]
    data = [header]
    for idx, o in enumerate(rows, start=1):
        data.append([
            _ar_shape(str(idx)), _pdf_cell(o["order_no"]),
            _pdf_cell(str(o["created_at"] or "")[:10]),
            _pdf_cell(o["reporter_name"]),
            _pdf_cell(o["location"]),
            _pdf_cell(o["section"]), _pdf_cell(o["technician"]),
            _pdf_cell(o["priority"]),
            _pdf_cell(STATUS_META.get(o["status"], (o["status"], "warn"))[0]),
        ])
    table = Table(data, colWidths=[9 * mm, 26 * mm, 22 * mm, 30 * mm, 42 * mm,
                                   26 * mm, 22 * mm, 20 * mm, 24 * mm],
                  repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        _pdf_escape(_ar_shape(f"{HOSPITAL_NAME} — {ORG_RIGHT} | Generated: "
                              f"{now_local().strftime('%Y/%m/%d')}")),
        ParagraphStyle("ft", fontName=_FONT_NAME, fontSize=7, leading=10,
                       alignment=TA_CENTER, textColor=colors.HexColor("#64748b")),
    ))

    doc.build(story)
    return buf.getvalue()


def monthly_xlsx_bytes(rows, totals, year, month, by_section) -> bytes:
    """بناء مصنّف Excel للتقرير الشهري وإرجاع وحدات البايت."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير الشهري"

    fill = PatternFill("solid", fgColor="0D9488")
    white = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total = totals.get("total", len(rows))
    rate = totals.get("rate", 0)

    ws.merge_cells("A1:I1")
    ws["A1"] = f"{HOSPITAL_NAME} — التقرير الشهري {MONTHS_AR[month - 1]} {year}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = f"{ORG_RIGHT} · {DEPARTMENT_NAME}"
    ws["A2"].alignment = center
    ws["A2"].font = Font(size=9, color="64748B")

    r = 4
    kpi_headers = ["إجمالي الأوامر", "منجزة", "قيد التنفيذ", "بانتظار الاعتماد",
                   "مرفوضة", "نسبة الإنجاز"]
    kpi_vals = [total, totals.get("done", 0), totals.get("approved", 0),
                totals.get("pending", 0), totals.get("rejected", 0), f"{rate}%"]
    for i, h in enumerate(kpi_headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = fill
        c.font = white
        c.alignment = center
        c.border = border
    for i, v in enumerate(kpi_vals, start=1):
        c = ws.cell(row=r + 1, column=i, value=v)
        c.alignment = center
        c.border = border
        c.font = Font(bold=True, size=12)

    r += 3
    ws.cell(row=r, column=1, value="توزيع البلاغات حسب القسم").font = Font(bold=True)
    r += 1
    for i, h in enumerate(["القسم / Type", "العدد", "النسبة"], start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = fill
        c.font = white
        c.alignment = center
        c.border = border
    r += 1
    for k, v in sorted(by_section.items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=1, value=k).border = border
        ws.cell(row=r, column=2, value=v).border = border
        pct = ws.cell(row=r, column=3, value=f"{round(v * 100 / total) if total else 0}%")
        pct.border = border
        pct.alignment = center
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="سجل أوامر العمل التفصيلي").font = Font(bold=True)
    r += 1
    headers = ["م", "رقم الأمر", "التاريخ", "المبلّغ", "الموقع", "القسم",
               "الفني", "الأولوية", "الحالة"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = fill
        c.font = white
        c.alignment = center
        c.border = border
    r += 1
    for idx, o in enumerate(rows, start=1):
        status_txt = STATUS_META.get(o["status"], (o["status"], "warn"))[0]
        vals = [idx, o["order_no"], str(o["created_at"] or "")[:10],
                o["reporter_name"], o["location"], o["section"],
                o["technician"] or "—", o["priority"], status_txt]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(i == 5))
        r += 1

    widths = [6, 18, 12, 22, 32, 22, 18, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------------------------------------------------- #
# إشعارات البريد (SMTP) و WhatsApp — كلها اختيارية وتُتجاهل إن لم تُضبط
# ----------------------------------------------------------------------- #
def _env(key: str) -> str:
    return os.environ.get(key, "").strip()


def notify_new_request(order_no: str, section: str, location: str,
                       reporter_name: str, contact: str, description: str):
    """إرسال تنبيه فوري (بريد و/أو WhatsApp) عند وصول بلاغ QR جديد."""
    _notify_email(order_no, section, location, reporter_name, contact, description)
    _notify_whatsapp(order_no, section, location, reporter_name, contact, description)


def _notify_email(order_no, section, location, reporter_name, contact, description):
    if not (_env("SMTP_USER") and _env("SMTP_PASSWORD") and _env("NOTIFY_TO")):
        return
    to = _env("NOTIFY_TO")
    subject = f"[بلاغ صيانة جديد] {order_no} - {section}"
    body = (
        f"بلاغ صيانة جديد 🛠️\n\n"
        f"رقم الأمر: {order_no}\n"
        f"الموقع: {location}\n"
        f"القسم: {section}\n"
        f"المُبلّغ: {reporter_name} ({contact})\n\n"
        f"الوصف:\n{description}\n"
    )
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = _env("SMTP_USER")
    msg["To"] = to
    msg.attach(MIMEText(body, "plain", "utf-8"))
    try:
        server = smtplib.SMTP(_env("SMTP_HOST") or "smtp.gmail.com",
                              int(_env("SMTP_PORT") or 587), timeout=20)
        server.starttls()
        server.login(_env("SMTP_USER"), _env("SMTP_PASSWORD"))
        server.sendmail(_env("SMTP_USER"), [to], msg.as_string())
        server.quit()
    except Exception:
        pass


def _notify_whatsapp(order_no, section, location, reporter_name, contact, description):
    if not (_env("WHATSAPP_TOKEN") and _env("WHATSAPP_PHONE_ID") and _env("WHATSAPP_TO")):
        return
    body = (
        f"بلاغ صيانة جديد 🛠️\n\n"
        f"رقم الأمر: {order_no}\n"
        f"الموقع: {location}\n"
        f"القسم: {section}\n"
        f"المُبلّغ: {reporter_name} ({contact})\n"
        f"الوصف: {description[:150]}\n"
    )
    payload = {
        "messaging_product": "whatsapp",
        "to": _env("WHATSAPP_TO"),
        "type": "text",
        "text": {"body": body},
    }
    url = f"https://graph.facebook.com/v19.0/{_env('WHATSAPP_PHONE_ID')}/messages"
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"Bearer {_env('WHATSAPP_TOKEN')}",
                 "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            resp.read()
    except Exception:
        pass


def insert_order(reporter_name, contact, location, section, priority,
                 technician, description, problem_type="", building="",
                 unit="", media="", status="pending", source="manual") -> tuple[int, str, str]:
    if priority not in PRIORITIES:
        priority = "عادية"
    now = now_local()
    with db() as c:
        cur = c.execute(
            """
            INSERT INTO work_orders
                (reporter_name, contact, location, building, unit, section,
                 problem_type, priority, technician, description, media,
                 status, source, token, created_at, completed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                reporter_name.strip(), contact.strip(), location.strip(),
                building.strip(), unit.strip(), section.strip() or "أخرى",
                problem_type.strip(), priority, technician.strip(),
                description.strip(), media, status, source,
                secrets.token_hex(8), fmt(now),
                fmt(now) if status == "done" and source == "manual" else None,
            ),
        )
        oid = cur.lastrowid
        order_no = f"WO-{now:%Y%m}-{oid:04d}"
        reference_id = f"R{now:%Y%m%d}-{secrets.token_hex(3).upper()}"
        c.execute(
            "UPDATE work_orders SET order_no=?, reference_id=? WHERE id=?",
            (order_no, reference_id, oid),
        )
    return oid, order_no, reference_id


def get_order_or_404(request: Request, order_id: int):
    with db() as c:
        row = c.execute("SELECT * FROM work_orders WHERE id=?", (order_id,)).fetchone()
    if row is None:
        return None
    return row


# --------------------------------------------------------------------------- #
# لوحة التحكم الرئيسية
# --------------------------------------------------------------------------- #
@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, status: str = "", q: str = ""):
    sql = "SELECT * FROM work_orders"
    conds, params = [], []
    if status in STATUS_META:
        conds.append("status=?")
        params.append(status)
    if q.strip():
        like = f"%{q.strip()}%"
        conds.append(
            "(reporter_name LIKE ? OR location LIKE ? OR order_no LIKE ? "
            "OR description LIKE ? OR technician LIKE ? OR contact LIKE ?)"
        )
        params.extend([like] * 6)
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT 500"

    with db() as c:
        orders = c.execute(sql, params).fetchall()
        stats = c.execute(
            """
            SELECT
                COUNT(*)                                     AS total,
                COUNT(CASE WHEN status='pending'  THEN 1 END) AS pending,
                COUNT(CASE WHEN status='approved' THEN 1 END) AS approved,
                COUNT(CASE WHEN status='done'     THEN 1 END) AS done,
                COUNT(CASE WHEN status='rejected' THEN 1 END) AS rejected,
                COUNT(CASE WHEN priority='طارئة'  THEN 1 END) AS urgent
            FROM work_orders
            """
        ).fetchone()
        dist_status, dist_section, dist_tech = {}, {}, {}
        dist_priority = {}
        for r in c.execute("SELECT status, section, technician, priority FROM work_orders").fetchall():
            dist_status[r["status"]] = dist_status.get(r["status"], 0) + 1
            s = r["section"] or "غير محدد"
            dist_section[s] = dist_section.get(s, 0) + 1
            t = (r["technician"] or "غير معيّن").strip() or "غير معيّن"
            dist_tech[t] = dist_tech.get(t, 0) + 1
            dist_priority[r["priority"]] = dist_priority.get(r["priority"], 0) + 1

    chart = {
        "status": dist_status,
        "section": dict(sorted(dist_section.items(), key=lambda kv: -kv[1])),
        "tech": dict(sorted(dist_tech.items(), key=lambda kv: -kv[1])),
        "priority": dist_priority,
    }

    return templates.TemplateResponse(
        request, "dashboard.html",
        context(request, orders=orders, stats=stats, q=q, status=status,
                chart_data=json.dumps(chart, ensure_ascii=False)),
    )


# --------------------------------------------------------------------------- #
# إنشاء أمر عمل
# --------------------------------------------------------------------------- #
@app.get("/orders/new", response_class=HTMLResponse)
def new_order_page(request: Request):
    return templates.TemplateResponse(request, "new_order.html", context(request))


@app.post("/orders/create")
def create_order(
    reporter_name: str = Form(...),
    contact: str = Form(""),
    location: str = Form(""),
    building: str = Form(""),
    unit: str = Form(""),
    section: str = Form(""),
    problem_type: str = Form(""),
    priority: str = Form("عادية"),
    technician: str = Form(""),
    description: str = Form(""),
    media_files: list[UploadFile] = File(default=[]),
):
    if not reporter_name.strip() or not description.strip():
        return RedirectResponse("/orders/new?err=1", status_code=303)
    # بناء الموقع المفصل من المبنى + الوحدة أو الحقل اليدوي
    if not location.strip():
        location = " / ".join(p for p in [building, unit] if p.strip())
    saved = save_media(media_files)
    oid, _order_no, _ref = insert_order(
        reporter_name, contact, location, section, priority,
        technician, description,
        problem_type=problem_type, building=building, unit=unit,
        media=",".join(saved), status="approved", source="manual",
    )
    return RedirectResponse(f"/orders/{oid}?created=1", status_code=303)


# --------------------------------------------------------------------------- #
# تفاصيل أمر العمل
# --------------------------------------------------------------------------- #
@app.get("/orders/{order_id}", response_class=HTMLResponse)
def order_detail(order_id: int, request: Request):
    row = get_order_or_404(request, order_id)
    if row is None:
        return templates.TemplateResponse(request, "404.html", context(request), status_code=404)
    return templates.TemplateResponse(
        request, "order_detail.html",
        context(request, o=row, media=media_list(row["media"])),
    )


@app.get("/orders/{order_id}/edit", response_class=HTMLResponse)
def order_edit_page(order_id: int, request: Request):
    row = get_order_or_404(request, order_id)
    if row is None:
        return templates.TemplateResponse(request, "404.html", context(request), status_code=404)
    return templates.TemplateResponse(
        request, "order_edit.html",
        context(request, o=row, media=media_list(row["media"])),
    )


@app.post("/orders/{order_id}/update")
def update_order(order_id: int, status: str = Form(...), technician: str = Form(""),
                 notes: str = Form("")):
    if status not in STATUS_META:
        raise HTTPException(status_code=400, detail="invalid status")
    now = now_local()
    with db() as c:
        existing = c.execute(
            "SELECT status, completed_at FROM work_orders WHERE id=?", (order_id,)
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="not found")
        completed_at = existing["completed_at"]
        if status == "done" and existing["status"] != "done":
            completed_at = fmt(now)
        elif status != "done":
            completed_at = None
        c.execute(
            """
            UPDATE work_orders
            SET status=?, technician=?, notes=?,
                completed_at=CASE WHEN ?!='' THEN ? ELSE ? END
            WHERE id=?
            """,
            (
                status,
                technician.strip() or existing["technician"],
                notes.strip(),
                completed_at, completed_at, completed_at,
                order_id,
            ),
        )
    return RedirectResponse(f"/orders/{order_id}?updated=1", status_code=303)


@app.post("/orders/{order_id}/edit")
def order_edit_submit(
    order_id: int,
    reporter_name: str = Form(...),
    contact: str = Form(""),
    location: str = Form(""),
    building: str = Form(""),
    unit: str = Form(""),
    section: str = Form(""),
    problem_type: str = Form(""),
    priority: str = Form("عادية"),
    technician: str = Form(""),
    description: str = Form(""),
):
    if not reporter_name.strip():
        return RedirectResponse(f"/orders/{order_id}/edit?err=1", status_code=303)
    if not location.strip():
        location = " / ".join(p for p in [building, unit] if p.strip())
    with db() as c:
        c.execute(
            """
            UPDATE work_orders
            SET reporter_name=?, contact=?, location=?, building=?, unit=?,
                section=?, problem_type=?, priority=?, technician=?, description=?
            WHERE id=?
            """,
            (
                reporter_name.strip(), contact.strip(), location.strip(),
                building.strip(), unit.strip(), section.strip(),
                problem_type.strip(), priority, technician.strip(),
                description.strip(), order_id,
            ),
        )
    return RedirectResponse(f"/orders/{order_id}?updated=1", status_code=303)


@app.post("/orders/{order_id}/delete")
def delete_order(order_id: int):
    with db() as c:
        c.execute("DELETE FROM work_orders WHERE id=?", (order_id,))
    return RedirectResponse("/?deleted=1", status_code=303)


# --------------------------------------------------------------------------- #
# طباعة أمر العمل (بالهيدر الرسمي)
# --------------------------------------------------------------------------- #
@app.get("/orders/{order_id}/print", response_class=HTMLResponse)
def print_order(order_id: int, request: Request, autoprint: int = 0):
    row = get_order_or_404(request, order_id)
    if row is None:
        return templates.TemplateResponse(request, "404.html", context(request), status_code=404)
    return templates.TemplateResponse(
        request, "print_order.html",
        context(request, o=row, autoprint=bool(autoprint), media=media_list(row["media"])),
    )


# --------------------------------------------------------------------------- #
# الطلبات الواردة (تدقيق بلاغات QR)
# --------------------------------------------------------------------------- #
@app.get("/incoming", response_class=HTMLResponse)
def incoming(request: Request):
    with db() as c:
        rows = c.execute(
            "SELECT * FROM work_orders WHERE status='pending' ORDER BY id ASC"
        ).fetchall()
    return templates.TemplateResponse(request, "incoming.html", context(request, rows=rows))


@app.post("/incoming/{order_id}/approve")
def approve_request(order_id: int):
    with db() as c:
        c.execute(
            "UPDATE work_orders SET status='approved' WHERE id=? AND status='pending'",
            (order_id,),
        )
    return RedirectResponse(f"/orders/{order_id}?approved=1", status_code=303)


@app.post("/incoming/{order_id}/reject")
def reject_request(order_id: int):
    with db() as c:
        c.execute(
            "UPDATE work_orders SET status='rejected' WHERE id=? AND status='pending'",
            (order_id,),
        )
    return RedirectResponse("/incoming?rejected=1", status_code=303)


# --------------------------------------------------------------------------- #
# رموز QR
# --------------------------------------------------------------------------- #
@app.get("/qr/{order_id}")
def qr_for_order(order_id: int, request: Request):
    row = get_order_or_404(request, order_id)
    if row is None:
        return Response(status_code=404)
    return make_qr_png(f"{base_url(request)}/track/{row['token']}")


@app.get("/orders/{order_id}/qr.png")
def qr_download(order_id: int, request: Request):
    """تحميل رمز QR الخاص بأمر العمل كصورة PNG (لتثبيته عند الوحدة)."""
    row = get_order_or_404(request, order_id)
    if row is None:
        raise HTTPException(status_code=404, detail="not found")
    resp = make_qr_png(f"{base_url(request)}/track/{row['token']}")
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="qr-{row["order_no"] or order_id}.png"'
    )
    return resp


@app.get("/poster-qr")
def poster_qr(request: Request, location: str = ""):
    url = f"{base_url(request)}/report"
    if location:
        from urllib.parse import quote
        url += f"?location={quote(location)}"
    return make_qr_png(url)


@app.get("/poster", response_class=HTMLResponse)
def poster_page(request: Request, location: str = ""):
    return templates.TemplateResponse(request, "poster.html", context(request, location=location))


# --------------------------------------------------------------------------- #
# النموذج العام (يُفتح بمسح QR)
# --------------------------------------------------------------------------- #
@app.get("/report", response_class=HTMLResponse)
def public_report(request: Request, location: str = "", ok: int = 0, err: int = 0):
    return templates.TemplateResponse(
        request, "public_report.html",
        context(request, location=location, ok=ok, err=err),
    )


@app.post("/report", response_class=HTMLResponse)
def public_report_submit(
    request: Request,
    reporter_name: str = Form(...),
    contact: str = Form(""),
    location: str = Form(""),
    building: str = Form(""),
    unit: str = Form(""),
    section: str = Form(""),
    problem_type: str = Form(""),
    priority: str = Form("عادية"),
    description: str = Form(""),
    media_files: list[UploadFile] = File(default=[]),
):
    if not reporter_name.strip() or not description.strip():
        return RedirectResponse("/report?err=1", status_code=303)
    if not location.strip():
        location = " / ".join(p for p in [building, unit] if p.strip())
    saved = save_media(media_files)
    _oid, order_no, reference_id = insert_order(
        reporter_name, contact, location, section, priority,
        "", description,
        problem_type=problem_type, building=building, unit=unit,
        media=",".join(saved), status="pending", source="qr",
    )
    # تنبيه فوري (بريد و/أو WhatsApp) — يُتجاهل بصمت إن لم تُضبط المتغيرات
    try:
        notify_new_request(order_no, section, location, reporter_name, contact, description)
    except Exception:
        pass
    return templates.TemplateResponse(
        request, "thanks.html",
        context(request, order_no=order_no, reference_id=reference_id, media=saved),
    )


# --------------------------------------------------------------------------- #
# صفحة تتبع أمر العمل (تفتح بمسح QR)
# --------------------------------------------------------------------------- #
@app.get("/track/{token}", response_class=HTMLResponse)
def track(token: str, request: Request):
    with db() as c:
        row = c.execute(
            "SELECT * FROM work_orders WHERE token=?", (token,)
        ).fetchone()
    if row is None:
        return templates.TemplateResponse(request, "404.html", context(request), status_code=404)
    return templates.TemplateResponse(
        request, "track.html",
        context(request, o=row, media=media_list(row["media"])),
    )


# --------------------------------------------------------------------------- #
# خدمة الملفات المرفوعة (الميديا)
# --------------------------------------------------------------------------- #
@app.get("/media/{filename}")
def serve_media(filename: str):
    # منع تجاوز المسار (path traversal)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=404, detail="not found")
    path = UPLOAD_DIR / filename
    if not path.is_file():
        raise HTTPException(status_code=404, detail="not found")
    ext = path.suffix.lower()
    mime = ALLOWED_MEDIA.get(ext, "application/octet-stream")
    return Response(content=path.read_bytes(), media_type=mime)


# --------------------------------------------------------------------------- #
# التقرير الشهري القابل للطباعة
# --------------------------------------------------------------------------- #
@app.get("/monthly", response_class=HTMLResponse)
def monthly_report(request: Request, year: int = 0, month: int = 0):
    t = now_local()
    year = year or t.year
    month = min(max(month or t.month, 1), 12)
    ym = f"{year:04d}-{month:02d}"

    with db() as c:
        rows = c.execute(
            "SELECT * FROM work_orders WHERE substr(created_at,1,7)=? ORDER BY id ASC",
            (ym,),
        ).fetchall()
        years = [
            int(r[0])
            for r in c.execute(
                "SELECT DISTINCT substr(created_at,1,4) AS y FROM work_orders ORDER BY y"
            ).fetchall()
        ]
    if not years or years[-1] < t.year:
        years.append(t.year)

    by_section, by_priority, by_status, by_source = {}, {}, {}, {}
    total_urgent = 0
    for r in rows:
        s = r["section"] or "غير محدد"
        by_section[s] = by_section.get(s, 0) + 1
        by_priority[r["priority"]] = by_priority.get(r["priority"], 0) + 1
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
        by_source[r["source"]] = by_source.get(r["source"], 0) + 1
        if r["priority"] == "طارئة":
            total_urgent += 1

    totals = {
        "total": len(rows),
        "done": by_status.get("done", 0),
        "approved": by_status.get("approved", 0),
        "pending": by_status.get("pending", 0),
        "rejected": by_status.get("rejected", 0),
        "urgent": total_urgent,
        "rate": round(by_status.get("done", 0) * 100 / len(rows)) if rows else 0,
    }

    return templates.TemplateResponse(
        request, "monthly.html",
        context(
            request,
            rows=rows, year=year, month=month, years=sorted(years),
            by_section=by_section, by_priority=by_priority,
            by_status=by_status, by_source=by_source, totals=totals,
            chart_data=json.dumps({
                "status": by_status,
                "section": dict(sorted(by_section.items(), key=lambda kv: -kv[1])),
                "priority": by_priority,
                "source": by_source,
            }, ensure_ascii=False),
        ),
    )


def _monthly_export_data(year: int, month: int):
    """إعادة بيانات التقرير الشهري نفسها المستخدمة في صفحة /monthly."""
    t = now_local()
    year = year or t.year
    month = min(max(month or t.month, 1), 12)
    ym = f"{year:04d}-{month:02d}"
    with db() as c:
        rows = c.execute(
            "SELECT * FROM work_orders WHERE substr(created_at,1,7)=? ORDER BY id ASC",
            (ym,),
        ).fetchall()
    by_section = {}
    by_status = {}
    for r in rows:
        s = r["section"] or "غير محدد"
        by_section[s] = by_section.get(s, 0) + 1
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    totals = {
        "total": len(rows),
        "done": by_status.get("done", 0),
        "approved": by_status.get("approved", 0),
        "pending": by_status.get("pending", 0),
        "rejected": by_status.get("rejected", 0),
        "rate": round(by_status.get("done", 0) * 100 / len(rows)) if rows else 0,
    }
    return rows, totals, by_section


@app.get("/monthly/pdf")
def monthly_report_pdf(year: int = 0, month: int = 0):
    """تحميل التقرير الشهري كملف PDF."""
    t = now_local()
    year = year or t.year
    month = min(max(month or t.month, 1), 12)
    rows, totals, by_section = _monthly_export_data(year, month)
    pdf_bytes = monthly_pdf_bytes(rows, totals, year, month, by_section)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="monthly-report-{year}-{month:02d}.pdf"'
            )
        },
    )


@app.get("/monthly/excel")
def monthly_report_excel(year: int = 0, month: int = 0):
    """تحميل التقرير الشهري كملف Excel (.xlsx)."""
    t = now_local()
    year = year or t.year
    month = min(max(month or t.month, 1), 12)
    rows, totals, by_section = _monthly_export_data(year, month)
    xlsx_bytes = monthly_xlsx_bytes(rows, totals, year, month, by_section)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="monthly-report-{year}-{month:02d}.xlsx"'
            )
        },
    )


# --------------------------------------------------------------------------- #
# حالة الخادم
# --------------------------------------------------------------------------- #
@app.get("/healthz")
def healthz():
    return {"status": "ok", "time": fmt(now_local()), "orders_db": str(DB_PATH)}


# --------------------------------------------------------------------------- #
# نقطة الدخول
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=port,
        proxy_headers=True,
        forwarded_allow_ips="*",
    )
